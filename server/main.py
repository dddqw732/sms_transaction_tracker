from __future__ import annotations

import json
import os
import random
import re
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

import jwt
from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from passlib.context import CryptContext
from pydantic import BaseModel, Field
# telethon is optional — only needed for Telegram SMS forwarding
try:
    from telethon import TelegramClient, events
    TELETHON_AVAILABLE = True
except ImportError:
    TelegramClient = None  # type: ignore
    events = None  # type: ignore
    TELETHON_AVAILABLE = False

import database
from sms_parser import parse_sms

JWT_SECRET = os.environ.get("JWT_SECRET", "dev-change-me")
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = int(os.environ.get("JWT_EXPIRE_HOURS", "24"))

pwd_context = CryptContext(schemes=["pbkdf2_sha256", "bcrypt"], deprecated="auto")


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def create_access_token(company_id: int, company_code: str) -> str:
    payload = {
        "sub": str(company_id),
        "company_code": company_code,
        "exp": now_utc() + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": now_utc(),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError as exc:
        raise HTTPException(status_code=401, detail="Session expired") from exc
    except jwt.InvalidTokenError as exc:
        raise HTTPException(status_code=401, detail="Invalid token") from exc


def normalize_status(value: str) -> str:
    status = value.strip().lower()
    if status not in {"active", "suspended"}:
        raise HTTPException(status_code=400, detail="Status must be Active or Suspended")
    return status


def company_to_public(company: dict) -> dict:
    return {
        "id": company["id"],
        "company_name": company["company_name"],
        "company_code": company["company_code"],
        "city": company["city"],
        "subscription_plan": company["subscription_plan"],
        "status": company["status"],
        "created_at": company["created_at"],
    }


def company_code_prefix(company_name: str) -> str:
    letters = re.sub(r"[^A-Za-z0-9]", "", company_name.upper())
    prefix = (letters[:3] if letters else "CMP").ljust(3, "X")
    return prefix


def generate_unique_company_code(company_name: str) -> str:
    prefix = company_code_prefix(company_name)
    next_num = database.get_next_company_code_number(prefix)

    # Sequential first, random fallback to avoid race/collision edge cases.
    candidates = [f"{prefix}{next_num:03d}"]
    for _ in range(30):
        candidates.append(f"{prefix}{random.randint(100, 999)}")

    return candidates[0], candidates[1:]


class SignupPayload(BaseModel):
    company_name: str = Field(min_length=2)
    city: str = Field(min_length=2)
    telegram_api_id: str = Field(min_length=2)
    telegram_api_hash: str = Field(min_length=6)
    initial_subscription_plan: str = Field(min_length=2)
    password: str = Field(min_length=6)
    status: str = "Active"


class LoginPayload(BaseModel):
    company_code: str = Field(min_length=3)
    password: str = Field(min_length=1)


class TransactionPayload(BaseModel):
    amount: float
    currency: str
    sender: str
    sender_number: Optional[str] = None
    receiver: str
    receiver_number: Optional[str] = None
    provider: str
    transaction_id: Optional[str] = None
    timestamp: str
    balance: Optional[float] = None
    type: str
    raw_sms: str


class InvoicePayload(BaseModel):
    invoice_number: str
    customer_phone: str
    amount: float
    currency: str = "SLSH"
    description: Optional[str] = None


class ConnectionManager:
    def __init__(self):
        self.active_connections: dict[int, list[WebSocket]] = {}

    async def connect(self, company_id: int, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.setdefault(company_id, []).append(websocket)

    def disconnect(self, company_id: int, websocket: WebSocket):
        connections = self.active_connections.get(company_id, [])
        if websocket in connections:
            connections.remove(websocket)
        if not connections and company_id in self.active_connections:
            del self.active_connections[company_id]

    async def broadcast(self, company_id: int, message: str):
        connections = self.active_connections.get(company_id, [])
        for connection in list(connections):
            try:
                await connection.send_text(message)
            except Exception:
                self.disconnect(company_id, connection)


manager = ConnectionManager()
telethon_client = None
telethon_company_id: Optional[int] = None


def choose_telethon_company() -> Optional[dict]:
    company_code = (os.environ.get("TELETHON_COMPANY_CODE") or "").strip().upper()
    if company_code:
        try:
            company = database.get_company_by_code(company_code)
        except Exception:
            return None
        if company and company.get("telegram_api_id") and company.get("telegram_api_hash"):
            return company

    try:
        companies = database.list_companies_with_telegram_credentials()
    except Exception:
        return None

    for company in companies:
        if company.get("status") == "active":
            return company
    return None


def load_telethon_env() -> None:
    env_path = Path(".env.telethon")
    if env_path.exists():
        with open(env_path, "r", encoding="utf-8") as file:
            for line in file:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    os.environ[key] = value


async def process_and_broadcast(msg_text: str, company_id: int):
    if not msg_text.strip():
        return

    sender_number = "Forwarded SMS"
    sms_body = msg_text
    lines = msg_text.split("\n")
    if len(lines) > 1 and "from:" in lines[0].lower():
        sender_number = lines[0].split(":", 1)[-1].strip()
        sms_body = "\n".join(lines[1:]).strip()

    parsed = parse_sms(sms_body, sender_number)
    if not parsed:
        return

    txn_id = database.insert_transaction(
        company_id=company_id,
        amount=parsed["amount"],
        currency=parsed["currency"],
        sender=parsed["sender"],
        sender_number=parsed.get("sender_number"),
        receiver=parsed["receiver"],
        receiver_number=parsed.get("receiver_number"),
        provider=parsed["provider"],
        transaction_id=parsed.get("transaction_id"),
        timestamp=parsed["timestamp"],
        balance=parsed.get("balance"),
        type_=parsed["type"],
        raw_sms=parsed["raw_sms"],
    )
    if txn_id is None:
        return

    parsed["id"] = txn_id

    if parsed["type"] == "Received":
        customer_phone = parsed.get("sender_number") or parsed.get("sender", "")
        matching_invoice = database.find_matching_invoice(company_id, customer_phone, parsed["amount"])
        if matching_invoice:
            database.update_invoice_status(company_id, matching_invoice["id"], "paid", txn_id)
            parsed["matched_invoice"] = matching_invoice["id"]

    database.create_notification(company_id, txn_id)
    await manager.broadcast(company_id, json.dumps(parsed, default=str))


@asynccontextmanager
async def lifespan(app: FastAPI):
    global telethon_client, telethon_company_id

    try:
        database.init_db()
        print("[OK] Database initialized.")
    except Exception as e:
        print(f"[WARNING] Database init failed (Supabase may not be configured): {e}")

    load_telethon_env()

    company = None
    try:
        company = choose_telethon_company()
    except Exception as e:
        print(f"[WARNING] Could not query companies (Supabase may not be configured): {e}")

    if company:
        telethon_company_id = int(company["id"])

    api_id = os.environ.get("TELEGRAM_API_ID") or (str(company.get("telegram_api_id")) if company else None)
    api_hash = os.environ.get("TELEGRAM_API_HASH") or (str(company.get("telegram_api_hash")) if company else None)

    if TELETHON_AVAILABLE and api_id and api_hash and Path("telethon.session").exists() and telethon_company_id is not None:
        try:
            telethon_client = TelegramClient("telethon", int(api_id), api_hash)

            @telethon_client.on(events.NewMessage(incoming=True))
            async def handle_new_message(event):
                sender = getattr(event, 'chat', None)
                sender_name = getattr(sender, 'username', None) or getattr(sender, 'title', None) or str(sender)
                print(f"[Telegram] Message from '{sender_name}': {(event.raw_text or '')[:120]}")
                if sender_name and 'wize' in str(sender_name).lower():
                    await process_and_broadcast(event.raw_text or "", telethon_company_id)
                elif 'smforward' in str(sender_name).lower() or 'smsforward' in str(sender_name).lower():
                    await process_and_broadcast(event.raw_text or "", telethon_company_id)

            @telethon_client.on(events.NewMessage(incoming=True, chats="WizeSMSForwardBot"))
            async def handle_wize_message(event):
                await process_and_broadcast(event.raw_text or "", telethon_company_id)

            await telethon_client.start()
            print(f"[OK] Telethon running for company_id={telethon_company_id} (api_id={api_id})")
        except Exception as e:
            print(f"[WARNING] Telethon failed to start: {e}")
            telethon_client = None
    else:
        print(f"[WARNING] Telethon not started. available={TELETHON_AVAILABLE}, session={Path('telethon.session').exists()}, company_id={telethon_company_id}, api_id={'set' if api_id else 'missing'}")

    yield

    if telethon_client:
        try:
            await telethon_client.disconnect()
        except Exception:
            pass


app = FastAPI(title="SMS Transaction Tracker API", lifespan=lifespan)


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc) or "Internal Server Error"},
    )


async def get_current_company(authorization: Optional[str] = Header(default=None)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token")

    token = authorization.split(" ", 1)[1].strip()
    payload = decode_token(token)

    company_id = int(payload["sub"])
    company = database.get_company_by_id(company_id)
    if not company:
        raise HTTPException(status_code=401, detail="Company not found")
    if company["status"] != "active":
        raise HTTPException(status_code=403, detail="Company account is suspended")
    return company


@app.get("/api/health")
def health_check():
    return {"status": "ok"}


@app.post("/api/auth/signup")
def signup_company(payload: SignupPayload):
    status = normalize_status(payload.status)
    telegram_api_id = payload.telegram_api_id.strip()
    telegram_api_hash = payload.telegram_api_hash.strip()
    if not telegram_api_id or not telegram_api_hash:
        raise HTTPException(status_code=400, detail="App api_id and App api_hash are required")

    password_hash = pwd_context.hash(payload.password)

    primary, fallback_codes = generate_unique_company_code(payload.company_name)
    codes_to_try = [primary, *fallback_codes]

    for company_code in codes_to_try:
        try:
            company = database.create_company(
                company_name=payload.company_name.strip(),
                company_code=company_code,
                city=payload.city.strip(),
                subscription_plan=payload.initial_subscription_plan.strip(),
                password_hash=password_hash,
                status=status,
                telegram_api_id=telegram_api_id,
                telegram_api_hash=telegram_api_hash,
            )
            token = create_access_token(int(company["id"]), company["company_code"])
            return {
                "status": "success",
                "company": company_to_public(company),
                "company_code": company["company_code"],
                "access_token": token,
                "token_type": "bearer",
            }
        except Exception:
            continue

    raise HTTPException(status_code=409, detail="Could not generate a unique company code")


@app.post("/api/auth/login")
def login_company(payload: LoginPayload):
    company_code = payload.company_code.strip().upper()
    company = database.get_company_by_code(company_code)
    if not company:
        raise HTTPException(status_code=401, detail="Invalid company code or password")

    if not pwd_context.verify(payload.password, company["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid company code or password")

    if company["status"] != "active":
        raise HTTPException(status_code=403, detail="Company account is suspended")

    token = create_access_token(int(company["id"]), company["company_code"])
    return {
        "status": "success",
        "access_token": token,
        "token_type": "bearer",
        "company": company_to_public(company),
    }


@app.get("/api/auth/me")
def auth_me(company: dict = Depends(get_current_company)):
    return {"status": "success", "company": company_to_public(company)}


@app.post("/api/transactions")
async def create_transaction(payload: TransactionPayload, company: dict = Depends(get_current_company)):
    txn_id = database.insert_transaction(
        company_id=company["id"],
        amount=payload.amount,
        currency=payload.currency,
        sender=payload.sender,
        sender_number=payload.sender_number,
        receiver=payload.receiver,
        receiver_number=payload.receiver_number,
        provider=payload.provider,
        transaction_id=payload.transaction_id,
        timestamp=payload.timestamp,
        balance=payload.balance,
        type_=payload.type,
        raw_sms=payload.raw_sms,
    )
    if txn_id is None:
        raise HTTPException(status_code=400, detail="Duplicate transaction_id")

    txn_data = payload.model_dump()
    txn_data["id"] = txn_id

    if payload.type == "Received":
        customer_phone = payload.sender_number or payload.sender
        matching_invoice = database.find_matching_invoice(company["id"], customer_phone, payload.amount)
        if matching_invoice:
            database.update_invoice_status(company["id"], matching_invoice["id"], "paid", txn_id)
            txn_data["matched_invoice"] = matching_invoice["id"]

    database.create_notification(company["id"], txn_id)
    await manager.broadcast(company["id"], json.dumps(txn_data, default=str))
    return {"status": "success", "id": txn_id}


@app.get("/api/transactions")
def list_transactions(
    search: Optional[str] = None,
    type: Optional[str] = None,
    provider: Optional[str] = None,
    sort_by: str = "timestamp",
    sort_order: str = "desc",
    company: dict = Depends(get_current_company),
):
    return database.get_transactions(
        company_id=company["id"],
        search=search,
        type_=type,
        provider=provider,
        sort_by=sort_by,
        sort_order=sort_order,
    )


@app.delete("/api/transactions")
async def clear_transactions(company: dict = Depends(get_current_company)):
    deleted = database.delete_all_transactions(company["id"])
    await manager.broadcast(company["id"], json.dumps({"event": "cleared"}))
    return {"status": "success", "deleted": deleted}


@app.post("/api/invoices")
def create_invoice(payload: InvoicePayload, company: dict = Depends(get_current_company)):
    invoice_id = database.create_invoice(
        company_id=company["id"],
        invoice_number=payload.invoice_number,
        customer_phone=payload.customer_phone,
        amount=payload.amount,
        currency=payload.currency,
        description=payload.description,
    )
    return {"status": "success", "id": invoice_id}


@app.get("/api/invoices")
def list_invoices(company: dict = Depends(get_current_company)):
    return database.get_invoices(company["id"])


@app.delete("/api/invoices/{invoice_id}")
def delete_invoice(invoice_id: int, company: dict = Depends(get_current_company)):
    database.delete_invoice(company["id"], invoice_id)
    return {"status": "success"}


@app.put("/api/invoices/{invoice_id}/status")
def update_invoice_status(invoice_id: int, status: str, company: dict = Depends(get_current_company)):
    clean_status = status.strip().lower()
    if clean_status not in {"pending", "paid", "cancelled"}:
        raise HTTPException(status_code=400, detail="Invalid invoice status")
    database.update_invoice_status(company["id"], invoice_id, clean_status)
    return {"status": "success"}


@app.get("/api/notifications")
def get_notifications(company: dict = Depends(get_current_company)):
    return database.get_unread_notifications(company["id"])


@app.put("/api/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, company: dict = Depends(get_current_company)):
    database.mark_notification_as_read(company["id"], notification_id)
    return {"status": "success"}


@app.delete("/api/notifications")
def clear_notifications(company: dict = Depends(get_current_company)):
    database.delete_all_notifications(company["id"])
    return {"status": "success"}


@app.get("/api/export/excel")
def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    company: dict = Depends(get_current_company),
):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    transactions = database.get_transactions(company_id=company["id"])

    if start_date or end_date:
        filtered = []
        for txn in transactions:
            txn_date = datetime.fromisoformat(str(txn["timestamp"]).replace("Z", "+00:00")).date()
            if start_date:
                start = datetime.fromisoformat(start_date).date()
                if txn_date < start:
                    continue
            if end_date:
                end = datetime.fromisoformat(end_date).date()
                if txn_date > end:
                    continue
            filtered.append(txn)
        transactions = filtered

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Type", "Sender", "Receiver", "Amount", "Currency", "Provider", "Transaction ID"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    received_total = 0
    sent_total = 0
    for txn in transactions:
        date = datetime.fromisoformat(str(txn["timestamp"]).replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
        ws.append(
            [
                date,
                txn["type"],
                txn["sender"],
                txn["receiver"],
                txn["amount"],
                txn["currency"],
                txn["provider"],
                txn.get("transaction_id") or "",
            ]
        )
        if txn["type"] == "Received":
            received_total += txn["amount"]
        else:
            sent_total += txn["amount"]

    ws.append([])
    ws.append(["Summary"])
    ws.append(["Total Received", "", "", "", received_total, "SLSH"])
    ws.append(["Total Sent", "", "", "", sent_total, "SLSH"])
    ws.append(["Net Balance", "", "", "", received_total - sent_total, "SLSH"])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = max_length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions_summary.xlsx"},
    )


APK_DIR = Path(__file__).parent / "downloads"
BUILD_APK_DIR = Path(__file__).parent.parent / "app" / "app" / "build" / "outputs" / "apk" / "debug"


@app.api_route("/download/app", methods=["GET", "HEAD"])
async def download_app():
    APK_DIR.mkdir(exist_ok=True)
    apk_files = list(APK_DIR.glob("*.apk"))
    if BUILD_APK_DIR.exists():
        apk_files.extend(list(BUILD_APK_DIR.glob("*.apk")))
    if not apk_files:
        raise HTTPException(status_code=404, detail="APK not built yet.")
    apk_path = max(apk_files, key=lambda file_path: file_path.stat().st_mtime)
    return FileResponse(
        path=str(apk_path),
        filename="FinSMS.apk",
        media_type="application/vnd.android.package-archive",
    )


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, token: Optional[str] = Query(default=None)):
    if not token:
        await websocket.close(code=4401)
        return

    try:
        payload = decode_token(token)
        company_id = int(payload["sub"])
        company = database.get_company_by_id(company_id)
        if not company or company["status"] != "active":
            await websocket.close(code=4403)
            return
    except Exception:
        await websocket.close(code=4401)
        return

    await manager.connect(company_id, websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(company_id, websocket)


app.mount("/", StaticFiles(directory=str(Path(__file__).parent / "static"), html=True), name="static")
